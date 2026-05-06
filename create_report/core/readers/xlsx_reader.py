from datetime import datetime
from openpyxl import load_workbook
from create_report.core.readers.base_reader import BaseReader


class EtalonReader(BaseReader):
    def read(self)->list:
        wb=load_workbook(self.file_path)
        ws=wb.active

        number=[]
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            value = row[0]
            if value is not None:
                number.append(str(value).strip())
        return number

class ReportReader(BaseReader):
    def read(self)->list:
        wb = load_workbook(self.file_path)
        ws=wb.active

        start_time=self._parse_period(ws["A3"].value)
        records = []

        for row in ws.iter_rows(min_row=12, min_col=1, max_col=2, values_only=True):
            number = row[0]
            time_str = row[1]

            if number is None:
                break

            adjusted_time = self._adjust_time(time_str, start_time)

            records.append({
                "number":str(number).strip(),
                "time":adjusted_time
            })
        return records

    def _parse_period(self, cell_value:str):
        parts= cell_value.split(" ")
        start_str=parts[2]+" "+parts[3]
        start_time=datetime.strptime(start_str,"%d.%m.%Y %H:%M:%S")
        return start_time

    def _adjust_time(self,time_str:str, start_time: datetime):
        if time_str is None:
            return "00:00:00"
        record_time = datetime.strptime(str(time_str).strip(), "%Y-%m-%d %H:%M:%S.%f")
        delta = record_time - start_time

        # timedelta в формат ЧЧ:ММ:СС
        total_seconds = int(delta.total_seconds())
        hours = total_seconds//3600
        minutes = (total_seconds%3600)//60
        seconds = total_seconds % 60

        return f"{hours:02}:{minutes:02}:{seconds:02}"