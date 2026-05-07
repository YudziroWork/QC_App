from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from create_report.core.report.base_report import BaseReport
from create_report.config.settings import COLORS
from create_report.core.stats.adv_stats import calculate_advanced_statistics
import os
from openpyxl.drawing.image import Image

class StandardReport(BaseReport):

    def _auto_width(self, ws):
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # буква столбца (A, B, C...)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 4  # +4 для отступа

    def generate(self,results: list, stats: dict, output_path:str,report_format: str = "xlsx"):
        wb=Workbook()
        ws= wb.active
        ws.title="Отчёт"

        self._write_comparicon_table(ws, results)
        self._write_missed(ws, stats['missed'])
        self._write_stats(ws, stats)
        self._auto_width(ws)

        adv_stats=calculate_advanced_statistics(results)
        ws2 = wb.create_sheet("Расширенная статистика")
        self._write_advanced_sheet(ws2, adv_stats)


        if report_format == "pdf":
            ws3= wb.create_sheet("Фото проблемных")
            self._write_problem_sheet(ws3,results,output_path)

            wb.save(output_path)

    # ─────────────────────────────────────────
    # Таблица сравнения A:D
    # ─────────────────────────────────────────

    def _write_comparicon_table(self, ws, results):
        headers = ["Распознанные", "Эталон", "Расстояние", "Время"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col,value=header)
            self._style_header(cell)

        for row_idx, record in enumerate(results, start=2):
            values = [
                record["recognized"],
                record["best_match"] if record["best_match"] else "",
                record["distance"] if record["distance"] is not None else"",
                record["time"]
            ]
            fill = self._get_fill(record["color"])

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                self._style_cell(cell)

    # ─────────────────────────────────────────
    # Пропущенные номера F:F
    # ─────────────────────────────────────────

    def _write_missed(self, ws, missed: list):
        header = ws.cell(row=1, column=6, value="Пропущенные номера")
        self._style_header(header)

        for row_idx, number in enumerate(missed, start=1):
            cell=ws.cell(row=row_idx+1, column=6, value=number)
            self._style_cell(cell)

    # ─────────────────────────────────────────
    # Статистика H:I и K:L
    # ─────────────────────────────────────────

    def _write_stats(self, ws, stats: dict):
        recognition_data = [
            ("Статистика распознания", None),
            ("Всего распознано", stats["total"]),
            ("Полностью распознано", stats["fully"]),
            ("Частично распознано", stats["partial"]),
            ("Неверно распознано", stats["incorrect"]),
            ("Без номера", stats["without_number"]),
            ("Число повторов", stats["duplicates"]),
        ]

        detection_data = [
            ("Статистика детекции", None),
            ("Всего записей в эталоне", stats["etalon_total"]),
            ("Обнаружено номеров эталона", stats["etalon_found"]),
            ("Пропущенные номера", stats["missed_count"]),
        ]

        quality_data = [
            ("Качество", None),
            ("Полностью распознано (%)", self._pct(stats["fully"], stats["total"])),
            ("Частично распознано (%)", self._pct(stats["partial"], stats["total"])),
            ("Неверно распознано (%)", self._pct(stats["incorrect"], stats["total"])),
            ("Пропущено (%)", self._pct(stats["missed_count"], stats["etalon_total"])),
            ("Повторы (%)", self._pct(stats["duplicates"], stats["total"])),
            ("Strict", None),
            ("Точность (%)", stats["precision_strict"]),
            ("Полнота (%)", stats["recall_strict"]),
            ("Конечная оценка (%)", stats["f1_strict"]),
            ("Soft", None),
            ("Точность (%)", stats["precision_soft"]),
            ("Полнота (%)", stats["recall_soft"]),
            ("Конечная оценка (%)", stats["f1_soft"]),
        ]

        self._write_block(ws, recognition_data, start_row=1, col_label=8, col_value=9)
        detection_start = 1 + len(recognition_data) + 2
        self._write_block(ws, detection_data, start_row=detection_start, col_label=8, col_value=9)
        self._write_block(ws, quality_data, start_row=1, col_label=11, col_value=12)

    def _write_block(self,ws,data: list, start_row: int, col_label: int,col_value: int):
        for i, (label, value) in enumerate(data):
            row=start_row + i
            label_cell = ws.cell(row=row, column=col_label, value=label)
            self._style_cell(label_cell)

            if value is not None:
                value_cell =ws.cell(row=row, column=col_value, value=value)
                self._style_cell(value_cell)

    def _write_advanced_sheet(self,ws,adv_stats: dict):
        ws.cell(row=1, column=1, value="Эталон")
        ws.cell(row=1, column=2, value="Повторы")
        self._style_header(ws.cell(row=1,column=1))
        self._style_header(ws.cell(row=1,column=2))
        ws.cell(row=1, column=4, value="Ошибка")
        ws.cell(row=1, column=5, value="Количество")
        self._style_header(ws.cell(row=1, column=4))
        self._style_header(ws.cell(row=1, column=5))

        row =2
        for etalon, count in sorted(adv_stats["etalon_usage"].items(),key=lambda x: x[1],reverse=True):
            ws.cell(row=row, column=1,value=etalon)
            ws.cell(row=row, column=2,value=count)
            self._style_cell(ws.cell(row=row,column=2))
            self._style_cell(ws.cell(row=row,column=2))
            row+=1

        row=2
        for error, count in sorted(adv_stats["error_stats"].items(),key=lambda x: x[1],reverse=True):
            ws.cell(row=row,column=4,value=error)
            ws.cell(row=row, column=5,value=count)
            self._style_cell(ws.cell(row=row, column=4))
            self._style_cell(ws.cell(row=row, column=5))
            row+=1

        self._auto_width(ws)

    # ─────────────────────────────────────────
    # 3-й лист
    # ─────────────────────────────────────────

    def _write_problem_sheet(self, ws, results: list, output_path: str):
        img_folder = os.path.join(os.path.dirname(output_path), "img")

        # заголовки
        ws.cell(row=1, column=1, value="Номер")
        ws.cell(row=1, column=2, value="Изображение")
        ws.cell(row=1, column=4, value="Нет номера")
        ws.cell(row=1, column=5, value="Изображение")
        for col in [1, 2, 4, 5]:
            self._style_header(ws.cell(row=1, column=col))

        if not os.path.exists(img_folder):
            return

        # список изображений "Нет номера"
        no_number_images = sorted(
            f for f in os.listdir(img_folder)
            if f.startswith("Нет номера")
        )
        no_number_index = 0
        problem_row = 2
        no_number_row = 2

        for item in results:
            recognized = item["recognized"]
            dist = item["distance"]

            # пропускаем зелёные
            if dist == 0:
                continue

            # блок "Нет номера"
            if item.get("no_number", False):
                ws.cell(row=no_number_row, column=4, value="Нет номера")

                if no_number_index < len(no_number_images):
                    img_path = os.path.join(img_folder, no_number_images[no_number_index])
                    self._insert_image(ws, img_path, f"E{no_number_row}", no_number_row)
                    no_number_index += 1

                no_number_row += 1
                continue

            # жёлтые и красные
            ws.cell(row=problem_row, column=1, value=recognized)
            self._style_cell(ws.cell(row=problem_row, column=1))

            images_found = sorted(
                f for f in os.listdir(img_folder)
                if f.startswith(recognized)
            )

            if images_found:
                img_path = os.path.join(img_folder, images_found[0])
                self._insert_image(ws, img_path, f"B{problem_row}", problem_row)

            problem_row += 1

    def _insert_image(self, ws, img_path: str, cell: str, row: int):
        try:
            img = Image(img_path)
            max_height = 250
            ratio = max_height / img.height
            img.height = max_height
            img.width = img.width * ratio
            ws.add_image(img, cell)
            ws.row_dimensions[row].height = max_height * 0.75
        except Exception as e:
            print(f"Ошибка вставки изображения {img_path}: {e}")



    # ─────────────────────────────────────────
    # Стили
    # ─────────────────────────────────────────
    def _style_header(self, cell):
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        self._apply_border(cell)

    def _style_cell(self, cell):
        cell.alignment= Alignment(horizontal="center")
        self._apply_border(cell)

    def _apply_border(self, cell):
        thin=Side(style="thin")
        cell.border=Border(left=thin, right=thin, top=thin, bottom=thin)

    def _get_fill(self,color: str)->PatternFill:
        hex_color=COLORS.get(color, "FFFFFF")
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

    def _pct(selfself, value: int,total: int)->float:
        if total == 0:
            return 0.0
        return round(value / total * 100, 2)

